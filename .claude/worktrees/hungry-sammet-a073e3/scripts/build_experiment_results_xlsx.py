"""Build docs/experiment_results.xlsx from a reproducible table layout."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "docs" / "experiment_results.xlsx"

EXPERIMENT_COLUMNS = [
    "Scenario/Domain",
    "Object amount",
    "Deadline",
    "Runs",
    "Search time",
    "Seed",
]

METRIC_COLUMNS = [
    "Completed",
    "Amount of success",
    "Average success time",
    "Std",
    "Percentage score",
    "Runtime",
]

ALGORITHMS = [
    "TP-MCTS itself (classical)",
    "parallel greedy classical PTRPG from TP-MCTS",
    "parallel greedy new PTRPG",
    "parallel greedy new PTRPG atomic propagation optimization",
]

EXPERIMENTS = [
    {"domain": "nasa_rover", "object_amount": 2, "deadline": 25, "runs": 20, "search_time": 1, "seed": 123},
    {"domain": "nasa_rover", "object_amount": 2, "deadline": 35, "runs": 20, "search_time": 1, "seed": 123},
    {"domain": "nasa_rover", "object_amount": 3, "deadline": 25, "runs": 20, "search_time": 1, "seed": 123},
    {"domain": "nasa_rover", "object_amount": 3, "deadline": 35, "runs": 20, "search_time": 1, "seed": 123},
]

# Prefill requested by user: map to Rover 2 settings row.
PREFILLED_RESULTS = [
    {
        "experiment_key": ("nasa_rover", 2, 25),
        "algorithm": "TP-MCTS itself (classical)",
        "metrics": {
            "Completed": 20,
            "Amount of success": 9,
            "Average success time": 23.444444444444443,
            "Std": 0.6260793149769247,
            "Percentage score": "45%",
            "Runtime": "32m 3.8s",
        },
    }
]

CONFIG_LINES = [
    "Settings description (demo.ipynb CONFIG baseline):",
    "",
    "CONFIG = {",
    '    "DOMAIN": "nasa_rover",',
    '    "OBJECT_AMOUNT": 2,',
    '    "GARBAGE_AMOUNT": 0,',
    '    "DEADLINE": 25,',
    '    "RUNS": 20,',
    '    "SEARCH_TIME": 1,',
    '    "SELECTION_TYPE": "avg",',
    '    "SEED": 123,',
    "}",
    "",
    "Across rover settings, change only OBJECT_AMOUNT and DEADLINE.",
]


def _cell_style(cell, *, fill=None, bold=False, size=None, h="center", v="center", wrap=True):
    if fill is not None:
        cell.fill = fill
    cell.font = Font(bold=bold, size=size)
    cell.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _set_vertical_border(cell, *, left=None, right=None):
    """Update only left/right border sides and keep existing top/bottom."""
    cell.border = Border(
        left=left if left is not None else cell.border.left,
        right=right if right is not None else cell.border.right,
        top=cell.border.top,
        bottom=cell.border.bottom,
    )


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    header_fill = PatternFill("solid", fgColor="D9E1F2")
    sub_fill = PatternFill("solid", fgColor="E7E6E6")

    # Left key columns: merged over header rows 1-2.
    for i, title in enumerate(EXPERIMENT_COLUMNS, start=1):
        ws.merge_cells(start_row=1, start_column=i, end_row=2, end_column=i)
        cell = ws.cell(row=1, column=i, value=title)
        _cell_style(cell, fill=header_fill, bold=True)

    # Algorithm groups on top row + metric names on second row.
    metric_count = len(METRIC_COLUMNS)
    start_col = len(EXPERIMENT_COLUMNS) + 1
    for algo in ALGORITHMS:
        end_col = start_col + metric_count - 1
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        algo_cell = ws.cell(row=1, column=start_col, value=algo)
        _cell_style(algo_cell, fill=header_fill, bold=True)
        for j, metric in enumerate(METRIC_COLUMNS):
            m_cell = ws.cell(row=2, column=start_col + j, value=metric)
            _cell_style(m_cell, fill=sub_fill, bold=True, size=10)
        start_col = end_col + 1

    # Data rows.
    first_data_row = 3
    experiment_row_map = {}
    for idx, exp in enumerate(EXPERIMENTS):
        row = first_data_row + idx
        ws.cell(row=row, column=1, value=exp["domain"])
        ws.cell(row=row, column=2, value=exp["object_amount"])
        ws.cell(row=row, column=3, value=exp["deadline"])
        ws.cell(row=row, column=4, value=exp["runs"])
        ws.cell(row=row, column=5, value=exp["search_time"])
        ws.cell(row=row, column=6, value=exp["seed"])
        experiment_row_map[(exp["domain"], exp["object_amount"], exp["deadline"])] = row

    # Fill known results by (experiment, algorithm).
    for entry in PREFILLED_RESULTS:
        row = experiment_row_map.get(entry["experiment_key"])
        if row is None:
            continue
        try:
            algo_idx = ALGORITHMS.index(entry["algorithm"])
        except ValueError:
            continue
        algo_start = len(EXPERIMENT_COLUMNS) + 1 + algo_idx * metric_count
        for j, metric in enumerate(METRIC_COLUMNS):
            ws.cell(row=row, column=algo_start + j, value=entry["metrics"].get(metric, ""))

    # Make algorithm group boundaries highly visible (header + subheader + data rows).
    thick_side = Side(style="thick")
    last_data_row = first_data_row + len(EXPERIMENTS) - 1
    group_start = len(EXPERIMENT_COLUMNS) + 1
    for _ in ALGORITHMS:
        group_end = group_start + metric_count - 1
        for row in range(1, last_data_row + 1):
            _set_vertical_border(ws.cell(row=row, column=group_start), left=thick_side)
            _set_vertical_border(ws.cell(row=row, column=group_end), right=thick_side)
        group_start = group_end + 1

    # Formatting and notes.
    ws.freeze_panes = "G3"
    widths = [18, 13, 10, 8, 12, 8]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    for c in range(len(EXPERIMENT_COLUMNS) + 1, len(EXPERIMENT_COLUMNS) + 1 + len(ALGORITHMS) * metric_count):
        ws.column_dimensions[get_column_letter(c)].width = 14

    note_row = first_data_row + len(EXPERIMENTS) + 2
    ws.cell(row=note_row, column=1, value="Experiment settings (baseline from demo.ipynb)").font = Font(
        bold=True, size=12
    )
    total_cols = len(EXPERIMENT_COLUMNS) + len(ALGORITHMS) * metric_count
    r = note_row + 1
    for line in CONFIG_LINES:
        ws.cell(row=r, column=1, value=line)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=total_cols)
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
